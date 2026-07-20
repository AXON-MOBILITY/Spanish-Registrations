#!/usr/bin/env python3
"""Extract a public-safe active BMW dealer list from an internal BUNO workbook."""

import argparse
import collections
import csv
import re
import unicodedata
from pathlib import Path


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = ROOT / "masters" / "master_bmw_active_dealers.csv"
FIELDS = (
    "dealer_id", "dealer_code", "dealer_name", "sales_installations",
    "source_kind", "source_confidence",
)


def normalize(value):
    value = unicodedata.normalize("NFKD", str(value or "").strip().upper())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    return " ".join(value.split())


def is_active_sales_installation(dist, group, installation_type):
    dist = normalize(dist)
    group = normalize(group)
    kind = normalize(installation_type)
    if dist == "ANT DEALER" or group in {"CERRADO", "TALLER AUTORIZADO"}:
        return False
    return (
        kind.startswith("PRIMERA INSTALACI")
        or "FULLY FLEDGED" in kind
        or "SALES SATELLITE" in kind
        or ("EXPOSICI" in kind and "V N" in kind)
    )


def slug(value):
    return normalize(value).lower().replace(" ", "-")


def extract_rows(workbook_path):
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["MASTER BUNO"]
    rows = sheet.iter_rows(values_only=True)
    headers = [normalize(value).replace(" ", "_") for value in next(rows)]
    grouped = collections.Counter()

    for values in rows:
        row = dict(zip(headers, values))
        if not is_active_sales_installation(
            row.get("DIST"), row.get("GRUPO"), row.get("TIPO_INSTALACION")
        ):
            continue
        raw_id = row.get("COD")
        dealer_id = str(int(raw_id)) if isinstance(raw_id, (int, float)) else str(raw_id or "").strip()
        dealer_name = str(row.get("DEALER_NAME") or "").strip()
        if not dealer_id or not dealer_name:
            continue
        grouped[(dealer_id, dealer_name)] += 1

    output = []
    names_by_code = collections.Counter(code for code, _ in grouped)
    for (dealer_code, dealer_name), installations in grouped.items():
        dealer_id = dealer_code
        if names_by_code[dealer_code] > 1:
            dealer_id = "{}:{}".format(dealer_code, slug(dealer_name))
        output.append({
            "dealer_id": dealer_id,
            "dealer_code": dealer_code,
            "dealer_name": dealer_name.lower(),
            "sales_installations": installations,
            "source_kind": "bmw_buno_internal",
            "source_confidence": "internal",
        })
    return sorted(output, key=lambda row: (int(row["dealer_code"]), row["dealer_id"]))


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()
    rows = extract_rows(args.workbook)
    if len(rows) != 54:
        raise RuntimeError("Expected 54 active BMW dealer names, found {}".format(len(rows)))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print("{} active BMW dealer names written to {}".format(len(rows), output))


if __name__ == "__main__":
    main()
