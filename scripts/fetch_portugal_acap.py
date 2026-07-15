# -*- coding: utf-8 -*-
"""fetch_portugal_acap.py — Ingesta mensual de matriculaciones de Portugal (ACAP).

Descarga los ficheros publicos que ACAP actualiza cada mes en URL fija:
  - XLSX "Matriculas de Veiculos Automoveis por Marca"
  - PDF  "Matriculas por tipo de energia" (solo se archiva)

Guarda snapshot fechado en data/portugal/raw/ y parsea el XLSX a
data/portugal/acap_marcas_snapshot_YYYYMMDD.csv + acumula en
data/portugal/pt_marcas_mensual.csv (serie por marca reconstruida por
diferencia de acumulados entre snapshots).

Uso:  python scripts/fetch_portugal_acap.py [--xlsx fichero_local.xlsx]
"""
import csv, datetime, io, os, re, sys, urllib.request

URL_XLSX = "https://www.acap.pt/site/uploads/paginas/documentos/07BAB4AD-CDBD0_1.xlsx"
URL_PDF_ENERGIA = "https://www.acap.pt/site/uploads/paginas/documentos/3E585470-43560_1.pdf"

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(_SCRIPT_DIR) if os.path.basename(_SCRIPT_DIR) == "scripts" else _SCRIPT_DIR
OUT_DIR = os.path.join(REPO_ROOT, "data", "portugal")
RAW_DIR = os.path.join(OUT_DIR, "raw")

MESES_PT = {m: i + 1 for i, m in enumerate(
    ["janeiro", "fevereiro", "marco", "abril", "maio", "junho",
     "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"])}


def _fetch(url, timeout=90):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def _norm(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").strip().lower()


def parse_xlsx(path):
    """Devuelve (periodo 'YYYY-MM', filas [{marca, mes_uds, acumulado_uds}]).

    Parser tolerante: busca la fila de cabecera que contenga 'marca' y un mes
    portugues; si el layout cambia, falla con mensaje claro (el raw ya queda
    archivado igualmente).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        header_idx = mes_col = acum_col = marca_col = None
        mes_num = year = None
        for i, row in enumerate(rows[:40]):
            cells = [_norm(c) for c in row]
            if any("marca" == c for c in cells):
                for j, c in enumerate(cells):
                    if c == "marca":
                        marca_col = j
                    if c in MESES_PT and mes_col is None:
                        mes_col, mes_num = j, MESES_PT[c]
                    if "janeiro" in c and "-" in c.replace(" ", "-"):
                        acum_col = j
                if marca_col is not None and mes_col is not None:
                    header_idx = i
                    break
        # anyo: buscar un 20xx en las primeras filas
        for row in rows[:10]:
            for c in row:
                m = re.search(r"20\d{2}", str(c or ""))
                if m:
                    year = int(m.group(0))
                    break
            if year:
                break
        if header_idx is None:
            continue
        out = []
        for row in rows[header_idx + 1:]:
            marca = str(row[marca_col] or "").strip()
            if not marca or _norm(marca) in ("total", "totais"):
                continue
            def num(j):
                try:
                    return int(float(row[j])) if j is not None and row[j] is not None else None
                except (TypeError, ValueError):
                    return None
            out.append({"marca": marca, "mes_uds": num(mes_col), "acum_uds": num(acum_col)})
        if out and year and mes_num:
            return f"{year}-{mes_num:02d}", out
    raise SystemExit("No se pudo parsear el XLSX de ACAP: revisar layout (raw archivado en data/portugal/raw/)")


def main():
    os.makedirs(RAW_DIR, exist_ok=True)
    today = datetime.date.today().strftime("%Y%m%d")

    local = None
    if "--xlsx" in sys.argv:
        local = sys.argv[sys.argv.index("--xlsx") + 1]

    xlsx_path = os.path.join(RAW_DIR, f"acap_marcas_{today}.xlsx")
    if local:
        import shutil
        shutil.copy(local, xlsx_path)
    else:
        open(xlsx_path, "wb").write(_fetch(URL_XLSX))
        try:
            open(os.path.join(RAW_DIR, f"acap_energia_{today}.pdf"), "wb").write(_fetch(URL_PDF_ENERGIA))
        except Exception as exc:
            print(f"WARN: PDF energia no descargado: {exc}")
    print(f"Snapshot raw: {xlsx_path}")

    periodo, filas = parse_xlsx(xlsx_path)
    snap_csv = os.path.join(OUT_DIR, f"acap_marcas_snapshot_{today}.csv")
    with open(snap_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["periodo", "marca", "uds_mes", "uds_acumulado"])
        for r in filas:
            w.writerow([periodo, r["marca"], r["mes_uds"], r["acum_uds"]])
    print(f"Parseado {len(filas)} marcas para {periodo} -> {snap_csv}")

    # Acumular serie mensual (mes = valor de columna mes del snapshot)
    serie = os.path.join(OUT_DIR, "pt_marcas_mensual.csv")
    existing = {}
    if os.path.exists(serie):
        with open(serie, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                existing[(row["periodo"], row["marca"])] = row["uds"]
    for r in filas:
        if r["mes_uds"] is not None:
            existing[(periodo, r["marca"])] = r["mes_uds"]
    with open(serie, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["periodo", "marca", "uds"])
        for (per, marca), uds in sorted(existing.items()):
            w.writerow([per, marca, uds])
    print(f"Serie mensual actualizada: {serie} ({len(existing)} filas)")


if __name__ == "__main__":
    main()
